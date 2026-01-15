import sys
import os


from mjira.const import FIELD_CSC, FIELD_SONY
import mjira.field
from mlogger import LOGGER
import os
import mjira.issue
import mjira.net_sony
import mzentao.trans
import mzentao.net
import mconfig
from openpyxl import load_workbook
from urllib.request import HTTPError
from utils import common
from libmirror.mstr import extract_pr_list


def download_jira_attachments(jira_info, jira_id):
    attachments = mjira.field.get_attachments(jira_info)
    assets_path_list = []
    assets_folder = os.getcwd() + "\\" + jira_id + "\\"
    for attachment_info in attachments:
        mjira.net_sony.download_sony_attachment(
            attachment_info["content"], attachment_info["filename"], assets_folder
        )
        assets_path_list.append(assets_folder + attachment_info["filename"])
    return assets_path_list


def sync_jira_to_zentao(jira_id, zt_pid,zt_assignee):
    try:
        # 获取Jira信息
        sync_result = ""
        input_bug_str = jira_id
        sony_jira_id = ""
        sync_success = "Success"
        sync_fail = "Fail"
        # 是否为DQA Jira Ticket
        input_key_str = mjira.issue.extract_internal_bug_key(input_bug_str)
        if not input_key_str:
            sync_result = jira_id + " - 输入值不为有效的 DQA Jira Ticket"
            return sync_fail, sync_result
        jira_info = mjira.net_sony.get_sony_jira(input_key_str)
        if not input_key_str:
            sync_result = jira_id + " - Sony Jira Center中搜索不到DQA Jira对应的信息"
            return sync_fail, sync_result
        sony_jira_id = input_key_str

        # 转换成Zt参数
        zt_id, zt_params = mzentao.trans.trans_sqa_to_zt_home(zt_pid, zt_assignee, jira_info)
        if zt_id:
            sync_result = jira_id + " - 此票已存在禅道ID: " + zt_id
            return sync_fail, sync_result

        # 获取附件并加载到本地
        assets_path_list = download_jira_attachments(jira_info, sony_jira_id)
        LOGGER.debug("附件列表:" + str(assets_path_list))

        # 创建Zt, 此时应获得Zt Id
        add_result, add_zt_id = mzentao.net.add_bug(zt_params, assets_path_list)
        if not add_result:
            sync_result = "禅道创建失败!"
            if not add_zt_id:
                sync_result = "禅道创建成功, 但获取创建的禅道Id失败!"
            return sync_fail, sync_result
        sync_result = "已创建的禅道：" + add_zt_id
        LOGGER.debug("创建的禅道Id: " + str(add_zt_id))

        chandao_link = mjira.issue.format_chandao_link("禅道-" + add_zt_id)
        update_result = mjira.net_sony.update_sony_jira(
            sony_jira_id, {FIELD_SONY.EXTERNAL_ISSUE_ID: chandao_link}
        )
        if update_result:
            sync_result = "禅道" + add_zt_id + "已更新到Jira的Trd Party Id字段"
            print("已将禅道链接更新到Jira的Trd Party Id字段")
            return sync_success, sync_result
        else:
            sync_result = sync_fail, "禅道" + add_zt_id + "更新到Jira失败!"
            return sync_fail, sync_result
    except Exception as e:
        print(e)
        return sync_fail, str(e)


def sync_excel():
    file_path = mconfig.get_issue_list_file_path()

    # 一份读取值
    wb_data = load_workbook(file_path, data_only=True)
    ws_data = wb_data.active

    # 一份读取公式
    wb = load_workbook(file_path)
    ws = wb.active

    # 处理
    for i, (row_data, row) in enumerate(zip(ws_data.iter_rows(min_row=2, max_col=7),
                                            ws.iter_rows(min_row=2, max_col=7)), start=2):
        jira_id = row_data[0].value  # 读取值
        zt_pid = str(row_data[2].value)  # 读取值
        zt_assignee = str(row_data[3].value) # 读取值

        if not jira_id:
            break

        sync_flg, sync_result = sync_jira_to_zentao(jira_id, zt_pid,zt_assignee)

        row[5].value = sync_flg
        row[6].value = sync_result

    # 保存（保留原始公式）
    wb.save(file_path)


def get_jira_sync_list(jql_text):
    jira_issue_list = []
    # 直接从mconfig获取
    # jql = mconfig.get_issue_list_jql()
    jql = jql_text

    try:
        search_result = mjira.net_sony.search_sony_jira_in_params(jql,mconfig.get_issue_field(),-1)
        #  核心判断：如果返回的是HTTPError异常对象 → 主动抛出，向上传递
        if isinstance(search_result, HTTPError):
            raise search_result
        #  只有返回正常数据，才执行取值和循环逻辑
        jql_result_list = search_result["issues"]
        for result in jql_result_list:
            issue_dict = dict()
            external_issue_link = result['fields']['customfield_17000']
            zentao_id = (
            mjira.issue.extract_chandao_key(external_issue_link.strip()) or ""
            if external_issue_link
            else ""
        )
            issue_dict.update(
                {'jira_key': result['key'],
                 'jira_summary': result['fields']['summary'],
                 'jira_status': result['fields']['status']['name'],
                 'jira_rank': result['fields']['customfield_31801']['value'],
                 'jira_comments':result['fields']['comment']['comments'],
                 'zentao_id':  zentao_id
                 })
            jira_issue_list.append(issue_dict)
        # print(jira_issue_list)

    except HTTPError as e:
        #  捕获HTTP异常，直接返回异常对象e，e内包含完整的 e.code + e.reason
        return e
    except Exception as e:
        #  兜底捕获所有其他异常，返回空列表，防止程序卡死
        return []

        #  正常返回数据列表
    return jira_issue_list


def get_jira_with_zentao(jira_issue_list, ui=None):
    if isinstance(jira_issue_list, HTTPError):
        raise jira_issue_list
    zentao_id_list = list()
    for jira_issue in jira_issue_list:
        zentao_id_list.append(jira_issue['zentao_id'])
    total_count = len(zentao_id_list)
    zentao_result_list = mzentao.net.get_bugs_with_ids(
        zentao_id_list,
        # 进度回调函数：实时更新UI气泡进度
        progress_callback=lambda curr, total: ui.run_in_main_thread(
            ui.show_progress_tooltip,
            f"共{total}条数据，解析中 {curr}/{total}"
        ) if ui else None
    )
    result_list = [{**d1, **d2} for d1, d2 in zip(jira_issue_list, zentao_result_list)]
    return result_list


def sync_zentao_history_to_jira(full_jira_zentao_data_list, ui=None):
    """
    同步禅道历史到Jira评论【完整版：进度条+精准统计+同步完成明细收集】
    同步完成后返回：统计结果 + 详细明细文案，用于弹窗展示精准状态
    :param full_jira_zentao_data_list: 完整的jira+zentao数据列表
    :param ui: UI实例，用于进度条回调
    :return: tuple (统计字典, 弹窗详情文案str)
    """
    total_count = len(full_jira_zentao_data_list)
    success_count = 0  # 至少1个ID新增成功的Jira数
    fail_count = 0  # 真实调用接口失败的Jira数
    skip_count = 0  # 无禅道历史/无JiraKey 跳过数
    no_sync_count = 0  # 有禅道历史、无需同步的Jira数

    # ========== 新增：核心收集变量 ==========
    detail_msg_list = []  # 收集每条Jira的处理详情，用于弹窗展示
    sync_id_count_dict = {}  # 记录每个Jira成功同步的历史ID数量 {JiraKey:同步条数}

    # 遍历每一条数据，按索引计数，用于进度条
    for curr_index, single_data in enumerate(full_jira_zentao_data_list, start=1):
        # 进度条实时更新
        if ui:
            ui.run_in_main_thread(
                ui.show_progress_tooltip,
                f"正在同步禅道历史到Jira [{curr_index}/{total_count}]，处理中..."
            )

        # 1. 基础数据获取
        zentao_history = single_data.get("zentao_history")
        jira_key = single_data.get("jira_key", "").strip()
        curr_jira_detail = f"{jira_key}"  # 初始化当前Jira的详情文本
        sync_id_count = 0  # 初始化当前Jira同步成功的ID数量

        # ========== 状态1：跳过 → jira无对应禅道/无JiraKey ==========
        if not zentao_history or not jira_key:
            skip_count += 1
            curr_jira_detail += " → 跳过（无对应禅道记录）"
            detail_msg_list.append(curr_jira_detail)
            print(f"ℹ️ [{curr_index}/{total_count}] {curr_jira_detail}")
            continue

        # 2. 获取Jira评论池，拼接所有评论内容
        jira_comments = single_data.get("jira_comments", [])
        jira_comment_text_pool = ""
        for comment in jira_comments:
            jira_comment_text_pool += comment.get("body", "") + " "

        # 初始化状态标记
        curr_has_success = False
        curr_has_fail = False
        curr_all_exist = True

        # 3. 遍历禅道历史ID，逐个校验+新增
        for history_id, history_content in zentao_history.items():
            if history_id not in jira_comment_text_pool:
                curr_all_exist = False
                comment_content = f"【禅道历史记录 ID:{history_id}】\n{common.transfer_single_zentao_history(history_content)}"
                # 调用评论方法
                add_result = mjira.net_sony.add_jira_comment(jira_key, comment_content)

                if add_result:
                    jira_comment_text_pool += history_id + " "
                    curr_has_success = True
                    sync_id_count += 1  # 计数+1：同步成功的ID数量
                    print(f"✅ [{curr_index}/{total_count}] {jira_key} → 禅道ID:{history_id} 新增评论成功")
                else:
                    curr_has_fail = True
                    print(f"❌ [{curr_index}/{total_count}] {jira_key} → 禅道ID:{history_id} 新增评论失败（接口调用失败）")
            else:
                print(f"ℹ️ [{curr_index}/{total_count}] {jira_key} → 禅道ID:{history_id} 已存在，无需新增")

        # ========== 按规则拼接当前Jira的最终状态文案 ==========
        if curr_has_success:
            success_count += 1
            sync_id_count_dict[jira_key] = sync_id_count
            curr_jira_detail += f" → 同步了 {sync_id_count} 条"
        elif curr_has_fail:
            fail_count += 1
            curr_jira_detail += " → 同步失败（接口调用异常）"
        elif curr_all_exist:
            no_sync_count += 1
            curr_jira_detail += " → 无需同步（已是最新）"

        # 把当前Jira的详情加入列表，用于弹窗
        detail_msg_list.append(curr_jira_detail)

    # 组装统计结果
    stat_result = {
        "success": success_count,
        "no_sync": no_sync_count,
        "skip": skip_count,
        "fail": fail_count
    }

    # ========== 组装弹窗的【详细明细文案】- 分行格式化，美观整洁 ==========
    detail_total_msg = "禅道历史同步Jira完成，详情如下：\n"
    detail_total_msg += "——————————————————\n"
    for msg in detail_msg_list:
        detail_total_msg += f"{msg}\n"
    detail_total_msg += "——————————————————\n"
    detail_total_msg += f"总计：{total_count}条 | 成功同步：{success_count}条 | 跳过：{skip_count}条 | 无需同步：{no_sync_count}条 | 失败：{fail_count}条"

    # 返回：统计字典 + 弹窗详情文案
    return stat_result, detail_total_msg